"""
Utilitários para geração e leitura de planilhas Excel padronizadas.
"""
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


COLUNAS_PLANILHA = [
    ('numero_processo', 'Número do Processo'),
    ('nome_beneficiario', 'Nome do Beneficiário'),
    ('cpf', 'CPF'),
    ('matricula', 'Matrícula'),
    ('municipio', 'Município'),
    ('cargo', 'Cargo/Função'),
    ('tipo_beneficio', 'Tipo de Benefício'),
    ('data_concessao', 'Data de Concessão'),
    ('data_publicacao', 'Data de Publicação'),
    ('regra_aplicada', 'Regra Aplicada'),
    ('base_calculo', 'Base de Cálculo (R$)'),
    ('valor_concedido', 'Valor Concedido (R$)'),
    ('valor_pago_folha', 'Valor Pago na Folha (R$)'),
    ('tempo_contribuicao', 'Tempo de Contribuição'),
    ('tempo_servico_publico', 'Tempo de Serviço Público'),
    ('tempo_carreira', 'Tempo de Carreira'),
    ('tempo_no_cargo', 'Tempo no Cargo'),
    ('marco_temporal_ingresso', 'Marco Temporal de Ingresso'),
    ('media_contribuicoes', 'Média das Contribuições (R$)'),
    ('proporcionalidade_percentual', 'Proporcionalidade (%)'),
    ('idade_concessao', 'Idade na Concessão'),
    ('observacoes', 'Observações'),
]

TIPOS_BENEFICIO_VALIDOS = [
    'APOS_VOLUNTARIA',
    'APOS_VOLUNTARIA_PROP',
    'APOS_INCAPACIDADE',
    'APOS_COMPULSORIA',
    'PENSAO_MORTE',
    'REVISAO_REENQUADRAMENTO',
]


def gerar_planilha_padronizada(dados_lista):
    """
    Gera uma planilha Excel padronizada a partir de lista de dicionários.
    Retorna bytes da planilha.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Processos Extraídos'

    # Estilo do cabeçalho
    header_fill = PatternFill(start_color='1A3A5C', end_color='1A3A5C', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # Escreve cabeçalho
    for col_idx, (campo, label) in enumerate(COLUNAS_PLANILHA, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Aba de referência para tipos de benefício
    ws_ref = wb.create_sheet('Referência')
    ws_ref['A1'] = 'Tipos de Benefício Válidos'
    ws_ref['A1'].font = Font(bold=True)
    for i, tipo in enumerate(TIPOS_BENEFICIO_VALIDOS, start=2):
        ws_ref[f'A{i}'] = tipo

    # Escreve dados
    row_fill_even = PatternFill(start_color='F0F4F8', end_color='F0F4F8', fill_type='solid')
    for row_idx, dados in enumerate(dados_lista, start=2):
        fill = row_fill_even if row_idx % 2 == 0 else None
        for col_idx, (campo, _) in enumerate(COLUNAS_PLANILHA, start=1):
            valor = dados.get(campo, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=str(valor) if valor else '')
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=False)
            if fill:
                cell.fill = fill

    # Ajusta largura das colunas
    larguras = [20, 35, 16, 15, 20, 25, 25, 15, 15, 35, 15, 15, 15, 20, 20, 12, 40]
    for col_idx, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = largura

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'

    # Adiciona instrução
    ws_ref['C1'] = 'Instruções'
    ws_ref['C1'].font = Font(bold=True)
    instrucoes = [
        '1. Preencha ou corrija os dados na aba "Processos Extraídos"',
        '2. O campo "tipo_beneficio" deve conter exatamente um dos valores da coluna A',
        '3. Datas devem estar no formato DD/MM/AAAA ou AAAA-MM-DD',
        '4. Valores monetários devem usar ponto como separador decimal (ex: 1234.56)',
        '5. Salve e reimporte a planilha no sistema',
    ]
    for i, instrucao in enumerate(instrucoes, start=2):
        ws_ref[f'C{i}'] = instrucao

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def ler_planilha_padronizada(arquivo):
    """
    Lê uma planilha Excel padronizada e retorna lista de dicionários.
    """
    try:
        wb = load_workbook(arquivo, data_only=True)
    except Exception as e:
        raise ValueError(f'Erro ao abrir planilha: {e}')

    ws = None
    for nome in ['Processos Extraídos', 'Sheet', 'Sheet1', wb.sheetnames[0]]:
        if nome in wb.sheetnames:
            ws = wb[nome]
            break

    if ws is None:
        raise ValueError('Aba de dados não encontrada na planilha.')

    # Detecta cabeçalho
    header_row = None
    campos_map = {label.lower(): campo for campo, label in COLUNAS_PLANILHA}
    campos_map.update({campo.lower(): campo for campo, _ in COLUNAS_PLANILHA})

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        if row[0] and str(row[0]).strip():
            header_row = row_idx
            headers = [str(c).strip().lower() if c else '' for c in row]
            break

    if header_row is None:
        raise ValueError('Cabeçalho não encontrado na planilha.')

    # Mapeia índices das colunas
    col_map = {}
    for col_idx, header in enumerate(headers):
        if header in campos_map:
            col_map[campos_map[header]] = col_idx

    resultados = []
    erros = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
        # Pula linhas vazias
        if all(c is None or str(c).strip() == '' for c in row):
            continue

        dados = {}
        for campo, _ in COLUNAS_PLANILHA:
            if campo in col_map:
                val = row[col_map[campo]]
                dados[campo] = str(val).strip() if val is not None else ''
            else:
                dados[campo] = ''

        # Validação básica
        if not dados.get('cpf'):
            erros.append({'linha': row_idx, 'mensagem': 'CPF obrigatório não informado.'})
            continue

        if not dados.get('nome_beneficiario'):
            erros.append({'linha': row_idx, 'mensagem': f'Nome do beneficiário não informado (CPF: {dados["cpf"]}).'})
            continue

        resultados.append({'linha': row_idx, 'dados': dados})

    return resultados, erros
